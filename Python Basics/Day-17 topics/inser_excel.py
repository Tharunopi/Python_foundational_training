import openpyxl

wb = path = r"C:\Stack overflow\Python_foundational_training\Python Basics\Day-17 topics\Book1.xlsx"

wb = openpyxl.load_workbook(path)
a_wb = wb.active

col = a_wb.max_column
row = 5
name = ["n-1", "n-2", "n-3", "n-4", "n-5"]
roll_num = [1, 2, 3, 4, 5]
print(row, col)

for i in range(2, row + 1):
    for j in range(1, col + 1):
        cell = a_wb.cell(i, j)
        if j == 2:
            cell.value = roll_num[i-1]
        else:
            cell.value = name[i-1]
wb.save(path)