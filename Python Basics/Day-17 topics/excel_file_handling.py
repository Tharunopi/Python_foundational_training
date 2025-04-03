import openpyxl

path = r"C:\Stack overflow\Python_foundational_training\Python Basics\Day-17 topics\Book1.xlsx"

wb = openpyxl.load_workbook(path)
a_wb = wb.active
col = a_wb.max_column
row = a_wb.max_row

print(row, col)

for i in range(1, row + 1):
    for j in range(1, col + 1):
        cell = a_wb.cell(i, j)
        print(cell.value)